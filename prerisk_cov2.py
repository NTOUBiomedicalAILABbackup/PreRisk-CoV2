#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PreRisk-CoV2: Pre-exposure Risk Assessment for SARS-CoV-2
==========================================================
Development and External Validation of a Pre-Exposure Protein Biomarker Panel
and Machine Learning Model for Predicting SARS-CoV-2 Infection Risk

Author: NTOU Biomedical AI LAB
GitHub: https://github.com/NTOUBiomedicalAILAB/PreRisk-CoV2
"""

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import argparse
import os
import sys
import time
import warnings
from math import sqrt
from sklearn import preprocessing
from sklearn.metrics import (confusion_matrix, roc_curve, roc_auc_score,
                             precision_recall_curve, average_precision_score)
from sklearn.model_selection import LeaveOneOut
from sklearn.neighbors import KNeighborsClassifier
from imblearn.over_sampling import SMOTE
import openpyxl
warnings.filterwarnings('ignore')


###############################################################################
# DEFAULT 7-PROTEIN PANEL
###############################################################################

DEFAULT_PANEL = ['MCP-3', 'LIF-R', 'TRANCE', 'FGF-23', 'NT-3', 'CXCL1', 'CXCL6']


def resolve_protein_indices(protein_id, panel=None):
    """
    Resolve protein names to 0-based indices in the protein_id list.
    Matching is case-insensitive and strips surrounding whitespace.
    """
    if panel is None:
        panel = DEFAULT_PANEL

    lookup = {name.strip().upper(): idx for idx, name in enumerate(protein_id)}
    indices = []
    missing = []

    for name in panel:
        key = name.strip().upper()
        if key in lookup:
            indices.append(lookup[key])
        else:
            missing.append(name)

    if missing:
        raise ValueError(
            f"[ERROR] The following proteins were NOT found in the CSV:\n"
            f"  {missing}\n"
            f"  Available proteins: {protein_id}"
        )

    return indices


###############################################################################
# BANNER & PROGRESS BAR
###############################################################################

def print_banner():
    banner = """
    ╔═══════════════════════════════════════════════════════════════╗
    ║                      PreRisk-CoV2                             ║
    ║         SARS-CoV-2 Pre-exposure Risk Assessment               ║
    ║          KNN-GA Protein Biomarker Framework                   ║
    ╚═══════════════════════════════════════════════════════════════╝
    """
    print(banner)


def print_progress_bar(iteration, total, prefix='Progress', suffix='Complete',
                       length=40, fill='█'):
    percent = ("{0:.1f}").format(100 * (iteration / float(total)))
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + '-' * (length - filled_length)
    sys.stdout.write(f'\r  {prefix} |{bar}| {percent}% {suffix}')
    sys.stdout.flush()
    if iteration == total:
        print()


###############################################################################
# DATA PROCESSING
# NOTE: Each CSV is normalized independently (fit_transform on its own data),
#       consistent with the original research code.
###############################################################################

def missing_counts(data):
    missing = data.isnull().sum()
    missing = missing[missing > 0]
    missing.sort_values(inplace=True)
    df_mc = pd.DataFrame({'ColumnName': missing.index,
                          'MissingCount': missing.values})
    df_mc['Percentage(%)'] = df_mc['MissingCount'].apply(
        lambda x: round(x / data.shape[0] * 100, 2))
    return df_mc


def data_processing(df):
    """
    Process a single CSV dataframe.
    MinMaxScaler is always fit on this df's own data (independent normalization).
    """
    sample_id  = df['sample ID'].values
    protein_id = df.columns.tolist()[2:94]

    df_proc = df.drop(['sample ID'], axis=1).copy()
    df_proc['PCR result'] = df_proc['PCR result'].map(
        {'Not': 0, 'Detected': 1}).astype(int)

    label    = df_proc['PCR result'].values
    features = df_proc.drop(['PCR result'], axis=1).values

    scaler   = preprocessing.MinMaxScaler(feature_range=(0, 1))
    features = scaler.fit_transform(features)

    return sample_id, protein_id, features, label


###############################################################################
# KNN BUILDER
###############################################################################

def build_knn_model(n_neighbors=5, leaf_size=30, algorithm='auto',
                    weights='uniform', p=2):
    return KNeighborsClassifier(
        n_neighbors=n_neighbors,
        leaf_size=leaf_size,
        algorithm=algorithm,
        weights=weights,
        p=p,
    )


###############################################################################
# ROC / PR CURVES
###############################################################################

def plot_roc_pr_curves(y_true, y_proba, save_path=None):
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))

    ax1.set_box_aspect(1)
    auc_val         = roc_auc_score(y_true, y_proba)
    fpr, tpr, _     = roc_curve(y_true, y_proba)
    ax1.plot(fpr, tpr, marker='.', label=' (AUROC = %0.3f)' % auc_val)
    ax1.plot([0, 1], [0, 1], color='red', linestyle='--')
    ax1.fill_between(fpr, tpr, color='gray', alpha=0.2)
    ax1.set_xlim([0.0, 1.0]); ax1.set_ylim([0.0, 1.0])
    ax1.set_title('ROC'); ax1.set_xlabel('False Positive Rate')
    ax1.set_ylabel('True Positive Rate'); ax1.legend(loc='lower right')

    ax2.set_box_aspect(1)
    precision_v, recall_v, _ = precision_recall_curve(y_true, y_proba)
    auprc_val = average_precision_score(y_true, y_proba)
    ax2.plot(recall_v, precision_v, marker='.', label=' (AUPRC = %0.3f)' % auprc_val)
    ax2.fill_between(recall_v, precision_v, color='gray', alpha=0.2)
    ax2.set_xlim([0.0, 1.0]); ax2.set_ylim([0.0, 1.0])
    ax2.set_title('PR'); ax2.set_xlabel('Recall', fontsize=14)
    ax2.set_ylabel('Precision', fontsize=14); ax2.legend(loc='lower right')

    plt.tight_layout()
    if save_path:
        os.makedirs(os.path.dirname(os.path.abspath(save_path)), exist_ok=True)
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close(fig)


###############################################################################
# EXCEL SAVE — INTERNAL
# data_save columns: 0=iter 1=acc 2=spec 3=sens 4=prec 5=auroc 6=auprc 7=mcc 8=f1
###############################################################################

def save_internal_results(data_save, loop, output_path, sheet_name):
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    try:
        wb = openpyxl.load_workbook(output_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    wb.create_sheet(sheet_name, 0)
    ws = wb.active

    ws.append(['Results of each Cross-Validation iteration:'])
    ws.append([''])
    ws.append([' ', 'Accuracy', 'Specificity', 'Sensitivity',
               'Precision', 'AUROC', 'AUPRC', 'MCC', 'F1_score'])

    for row in data_save:
        ws.append(row.tolist())

    ws.append([''])
    ws.append(['', 'Accuracy', 'Specificity', 'Sensitivity',
               'Precision', 'AUROC', 'AUPRC', 'MCC', 'F1_score'])

    ws.append([
        'Overall Mean',
        sum(data_save[:, 1]) / loop,
        sum(data_save[:, 2]) / loop,
        sum(data_save[:, 3]) / loop,
        sum(data_save[:, 4]) / loop,
        sum(data_save[:, 5]) / loop,
        sum(data_save[:, 6]) / loop,
        sum(data_save[:, 7]) / loop,
        sum(data_save[:, 8]) / loop,
    ])
    ws.append([
        'Std Dev',
        np.std(data_save[:, 1]), np.std(data_save[:, 2]),
        np.std(data_save[:, 3]), np.std(data_save[:, 4]),
        np.std(data_save[:, 5]), np.std(data_save[:, 6]),
        np.std(data_save[:, 7]), np.std(data_save[:, 8]),
    ])
    ws.append([
        'Formatted Data',
        f"{round(sum(data_save[:,1])/loop*100,2)} ± {round(np.std(data_save[:,1])*100,2)}",
        f"{round(sum(data_save[:,2])/loop*100,2)} ± {round(np.std(data_save[:,2])*100,2)}",
        f"{round(sum(data_save[:,3])/loop*100,2)} ± {round(np.std(data_save[:,3])*100,2)}",
        f"{round(sum(data_save[:,4])/loop*100,2)} ± {round(np.std(data_save[:,4])*100,2)}",
        f"{round(sum(data_save[:,5])/loop,4)} ± {round(np.std(data_save[:,5]),4)}",
        f"{round(sum(data_save[:,6])/loop,4)} ± {round(np.std(data_save[:,6]),4)}",
        f"{round(sum(data_save[:,7])/loop,4)} ± {round(np.std(data_save[:,7]),4)}",
        f"{round(sum(data_save[:,8])/loop*100,2)} ± {round(np.std(data_save[:,8])*100,2)}",
    ])

    wb.save(output_path)
    print(f'[INFO] Results saved to: {output_path}')


###############################################################################
# EXCEL SAVE — EXTERNAL
# data_save columns: 0=iter 1=acc 2=sens 3=prec 4=f1 5=auroc 6=auprc 7=mcc 8=spec
###############################################################################

def save_external_results(data_save, loop, output_path, sheet_name):
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    try:
        wb = openpyxl.load_workbook(output_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    wb.create_sheet(sheet_name, 0)
    ws = wb.active

    ws.append(['', 'Accuracy', 'Sensitivity', 'Precision',
               'F1_score', 'AUROC', 'AUPRC', 'MCC', 'Specificity'])
    ws.append([
        'Overall Mean',
        sum(data_save[:, 1]) / loop, sum(data_save[:, 2]) / loop,
        sum(data_save[:, 3]) / loop, sum(data_save[:, 4]) / loop,
        sum(data_save[:, 5]) / loop, sum(data_save[:, 6]) / loop,
        sum(data_save[:, 7]) / loop, sum(data_save[:, 8]) / loop,
    ])
    ws.append([
        'Std Dev',
        np.std(data_save[:, 1]), np.std(data_save[:, 2]),
        np.std(data_save[:, 3]), np.std(data_save[:, 4]),
        np.std(data_save[:, 5]), np.std(data_save[:, 6]),
        np.std(data_save[:, 7]), np.std(data_save[:, 8]),
    ])
    ws.append([
        'Formatted Data',
        f"{round(sum(data_save[:,1])/loop*100,2)} ± {round(np.std(data_save[:,1])*100,2)}",
        f"{round(sum(data_save[:,2])/loop,4)} ± {round(np.std(data_save[:,2]),4)}",
        f"{round(sum(data_save[:,3])/loop,4)} ± {round(np.std(data_save[:,3]),4)}",
        f"{round(sum(data_save[:,4])/loop,4)} ± {round(np.std(data_save[:,4]),4)}",
        f"{round(sum(data_save[:,5])/loop,4)} ± {round(np.std(data_save[:,5]),4)}",
        f"{round(sum(data_save[:,6])/loop,4)} ± {round(np.std(data_save[:,6]),4)}",
        f"{round(sum(data_save[:,7])/loop,4)} ± {round(np.std(data_save[:,7]),4)}",
        f"{round(sum(data_save[:,8])/loop,4)} ± {round(np.std(data_save[:,8]),4)}",
    ])

    ws.append(['']); ws.append(['']); ws.append([''])
    ws.append(['Results of each prediction iteration:'])
    ws.append([''])
    ws.append([' ', 'Accuracy', 'Sensitivity', 'Precision',
               'F1_score', 'AUROC', 'AUPRC', 'MCC', 'Specificity'])
    for row in data_save:
        ws.append(row.tolist())

    wb.save(output_path)
    print(f'[INFO] Results saved to: {output_path}')


###############################################################################
# 1. INTERNAL VALIDATION (LOOCV)
###############################################################################

def internal_validation(args):
    print("\n" + "=" * 70)
    print("INTERNAL VALIDATION MODE (LOOCV)")
    print("=" * 70)

    print(f'[INFO] Loading data from: {args.input}')
    df = pd.read_csv(args.input)
    sample_id, protein_id, features, label = data_processing(df)

    print(f'[INFO] Dataset shape        : {features.shape}')
    print(f'[INFO] Class distribution   : {np.bincount(label.astype(int))}')

    if args.protein_indices:
        protein_indices = args.protein_indices
        print(f'[INFO] Selected proteins ({len(protein_indices)}) [manual index]: '
              f'{[protein_id[i] for i in protein_indices]}')
    else:
        panel = args.protein_names if args.protein_names else DEFAULT_PANEL
        protein_indices = resolve_protein_indices(protein_id, panel)
        print(f'[INFO] Selected proteins ({len(protein_indices)}) [auto name-match]: '
              f'{[protein_id[i] for i in protein_indices]}')

    features_sel = features[:, protein_indices]
    loop         = args.n_iterations
    loo          = LeaveOneOut()
    data_save    = np.zeros((loop, 9), dtype=float)

    print(f'[INFO] Running LOOCV (n_iterations={loop}) ...')
    total_start = time.time()

    for i in range(loop):
        if not args.verbose:
            print_progress_bar(i + 1, loop, prefix='Training Progress')

        ans, pred, probs = [], [], []
        for train_idx, test_idx in loo.split(features_sel):
            X_train = features_sel[train_idx]
            X_test  = features_sel[test_idx]
            y_train = label[train_idx]
            y_test  = label[test_idx]

            if args.use_smote:
                min_count = min(int(sum(y_train == 0)), int(sum(y_train == 1)))
                if min_count >= 2:
                    k = min(5, min_count - 1)
                    try:
                        sm = SMOTE(k_neighbors=k, random_state=i)
                        X_train, y_train = sm.fit_resample(X_train, y_train)
                    except Exception:
                        pass

            model = build_knn_model(
                n_neighbors=args.n_neighbors, leaf_size=args.leaf_size,
                algorithm=args.algorithm, weights=args.weights, p=args.p,
            )
            model.fit(X_train, y_train)
            ans.append(y_test[0])
            pred.append(model.predict(X_test)[0])
            probs.append(model.predict_proba(X_test)[:, 1][0])

        ans   = np.array(ans,   dtype=int)
        pred  = np.array(pred,  dtype=int)
        probs = np.array(probs, dtype=float)

        cm = confusion_matrix(ans, pred)
        TP = cm[1, 1]; FN = cm[1, 0]; FP = cm[0, 1]; TN = cm[0, 0]

        accuracy    = (TP + TN) / (TP + FP + FN + TN)
        specificity = TN / (TN + FP)       if (TN + FP) > 0 else 0
        sensitivity = TP / (TP + FN)       if (TP + FN) > 0 else 0
        precision   = TP / (TP + FP)       if (TP + FP) > 0 else 0
        f1_score    = (2 / ((1 / precision) + (1 / sensitivity))
                       if precision > 0 and sensitivity > 0 else 0)
        auroc       = roc_auc_score(ans, probs)
        auprc       = average_precision_score(ans, probs)
        mcc_denom   = sqrt((TP+FP)*(TP+FN)*(TN+FP)*(TN+FN))
        mcc         = (TP*TN - FP*FN) / mcc_denom if mcc_denom > 0 else 0

        data_save[i] = [i+1, accuracy, specificity, sensitivity,
                        precision, auroc, auprc, mcc, f1_score]

        if args.verbose:
            print(f'\nIteration {i+1} | Acc: {accuracy:.4f} | AUROC: {auroc:.4f}')

        if args.plot_curves and i == loop - 1:
            plot_roc_pr_curves(
                ans, probs,
                save_path=os.path.join(args.output_dir, 'internal_roc_pr.png')
            )

    elapsed = time.time() - total_start
    print(f'[INFO] Total time: {elapsed:.2f} s')
    print('=' * 79)
    print(f'Averages after {loop} cross-validation iterations:')
    print(f'Accuracy    = {round(sum(data_save[:,1])/loop*100,2)} ± {round(np.std(data_save[:,1])*100,2)}')
    print(f'Specificity = {round(sum(data_save[:,2])/loop*100,2)} ± {round(np.std(data_save[:,2])*100,2)}')
    print(f'Recall      = {round(sum(data_save[:,3])/loop*100,2)} ± {round(np.std(data_save[:,3])*100,2)}')
    print(f'Precision   = {round(sum(data_save[:,4])/loop*100,2)} ± {round(np.std(data_save[:,4])*100,2)}')
    print(f'AUROC       = {round(sum(data_save[:,5])/loop,4)} ± {round(np.std(data_save[:,5]),4)}')
    print(f'AUPRC       = {round(sum(data_save[:,6])/loop,4)} ± {round(np.std(data_save[:,6]),4)}')
    print(f'MCC         = {round(sum(data_save[:,7])/loop,4)} ± {round(np.std(data_save[:,7]),4)}')
    print(f'F1_score    = {round(sum(data_save[:,8])/loop*100,2)} ± {round(np.std(data_save[:,8])*100,2)}')
    print('=' * 79)

    local_time = time.strftime('%Y%m%d_%H%M%S', time.localtime())
    save_internal_results(
        data_save, loop,
        os.path.join(args.output_dir, f'Internal_Validation_{local_time}.xlsx'),
        'Sheet1'
    )


###############################################################################
# 2. EXTERNAL VALIDATION
# NOTE: train and test are each normalized independently (fit_transform),
#       matching the original research code behaviour.
###############################################################################

def external_validation(args):
    print("\n" + "=" * 70)
    print("EXTERNAL VALIDATION MODE")
    print("=" * 70)

    print(f'[INFO] Loading training data from : {args.train_input}')
    train_df = pd.read_csv(args.train_input)
    _, protein_id, train_features, train_label = data_processing(train_df)

    print(f'[INFO] Loading test data from     : {args.test_input}')
    test_df = pd.read_csv(args.test_input)
    _, _, test_features, test_label = data_processing(test_df)   # independent normalize

    print(f'[INFO] Training set shape         : {train_features.shape}')
    print(f'[INFO] Test set shape             : {test_features.shape}')

    if args.protein_indices:
        protein_indices = args.protein_indices
        print(f'[INFO] Selected proteins ({len(protein_indices)}) [manual index]: '
              f'{[protein_id[i] for i in protein_indices]}')
    else:
        panel = args.protein_names if args.protein_names else DEFAULT_PANEL
        protein_indices = resolve_protein_indices(protein_id, panel)
        print(f'[INFO] Selected proteins ({len(protein_indices)}) [auto name-match]: '
              f'{[protein_id[i] for i in protein_indices]}')

    train_sel = train_features[:, protein_indices]
    test_sel  = test_features[:,  protein_indices]

    goodbad   = np.zeros(len(test_sel), dtype=float)
    loop      = args.n_iterations
    data_save = np.zeros((loop, 9), dtype=float)

    knn = build_knn_model(
        n_neighbors=args.n_neighbors, leaf_size=args.leaf_size,
        algorithm=args.algorithm, weights=args.weights, p=args.p,
    )

    print(f'[INFO] Running external validation (n_iterations={loop}) ...')
    total_start = time.time()

    for loop_count in range(loop):
        if not args.verbose:
            print_progress_bar(loop_count + 1, loop, prefix='Prediction Progress')

        if args.use_smote:
            sm = SMOTE(k_neighbors=5, random_state=loop_count)
            X_train, y_train = sm.fit_resample(train_sel, train_label)
        else:
            X_train, y_train = train_sel, train_label

        knn.fit(X_train, y_train)
        prediction = knn.predict(test_sel)
        probs      = knn.predict_proba(test_sel)[:, 1]

        if args.plot_curves and loop_count == loop - 1:
            plot_roc_pr_curves(
                test_label, probs,
                save_path=os.path.join(args.output_dir, 'external_roc_pr.png')
            )

        cm = confusion_matrix(test_label, prediction)
        TP = cm[1, 1]; FP = cm[0, 1]; FN = cm[1, 0]; TN = cm[0, 0]

        accuracy    = (TP + TN) / (TP + FP + FN + TN)
        sensitivity = TP / (TP + FN) if (TP + FN) > 0 else 0
        precision   = TP / (TP + FP) if (TP + FP) > 0 else 0
        f1_score    = (2 / ((1 / precision) + (1 / sensitivity))
                       if precision > 0 and sensitivity > 0 else 0)
        auroc       = roc_auc_score(test_label, probs)
        auprc       = average_precision_score(test_label, probs)
        specificity = TN / (TN + FP) if (TN + FP) > 0 else 0
        mcc_denom   = sqrt((TP+FP)*(TP+FN)*(TN+FP)*(TN+FN))
        mcc         = (TP*TN - FP*FN) / mcc_denom if mcc_denom > 0 else 0

        data_save[loop_count] = [loop_count+1, accuracy, sensitivity, precision,
                                  f1_score, auroc, auprc, mcc, specificity]

        if args.verbose:
            print(f'\nIteration {loop_count+1} | Acc: {accuracy:.4f} | AUROC: {auroc:.4f}')

        for idx in range(len(goodbad)):
            if test_label[idx] == prediction[idx]:
                goodbad[idx] += 1

    elapsed = time.time() - total_start
    print(f'[INFO] Total time: {elapsed:.2f} s')
    print('=' * 79)
    print(f'Averages after {loop} prediction iterations:')
    print(f'Accuracy    = {round(sum(data_save[:,1])/loop*100,2)} ± {round(np.std(data_save[:,1])*100,2)}')
    print(f'Sensitivity = {round(sum(data_save[:,2])/loop,4)} ± {round(np.std(data_save[:,2]),4)}')
    print(f'Precision   = {round(sum(data_save[:,3])/loop,4)} ± {round(np.std(data_save[:,3]),4)}')
    print(f'F1_score    = {round(sum(data_save[:,4])/loop,4)} ± {round(np.std(data_save[:,4]),4)}')
    print(f'AUROC       = {round(sum(data_save[:,5])/loop,4)} ± {round(np.std(data_save[:,5]),4)}')
    print(f'AUPRC       = {round(sum(data_save[:,6])/loop,4)} ± {round(np.std(data_save[:,6]),4)}')
    print(f'MCC         = {round(sum(data_save[:,7])/loop,4)} ± {round(np.std(data_save[:,7]),4)}')
    print(f'Specificity = {round(sum(data_save[:,8])/loop,4)} ± {round(np.std(data_save[:,8]),4)}')
    print('=' * 79)

    local_time = time.strftime('%Y%m%d_%H%M%S', time.localtime())
    save_external_results(
        data_save, loop,
        os.path.join(args.output_dir, f'External_Validation_Results_{local_time}.xlsx'),
        'Sheet1'
    )


###############################################################################
# MAIN ENTRY
###############################################################################

def main():
    parser = argparse.ArgumentParser(
        description='PreRisk-CoV2: SARS-CoV-2 Pre-exposure Risk Assessment Framework',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Default 7-protein panel (auto name-match):
  {DEFAULT_PANEL}

Examples:
  # Internal validation — default 7-protein panel
  python prerisk_cov2.py --mode internal --input Discovery.csv --n-iterations 100 --use-smote --plot-curves

  # Internal validation — custom protein names
  python prerisk_cov2.py --mode internal --input Discovery.csv --protein-names MCP-3 LIF-R TRANCE

  # External validation
  python prerisk_cov2.py --mode external --train-input Discovery.csv --test-input Validation.csv --n-iterations 100 --use-smote --plot-curves
        """
    )

    parser.add_argument('--mode', type=str, required=True,
                        choices=['internal', 'external'],
                        help='Validation mode: internal (LOOCV) / external')

    parser.add_argument('--input',       type=str, help='Input CSV (internal mode)')
    parser.add_argument('--train-input', type=str, help='Training CSV (external mode)')
    parser.add_argument('--test-input',  type=str, help='Test CSV (external mode)')

    feat_grp = parser.add_mutually_exclusive_group()
    feat_grp.add_argument(
        '--protein-names', type=str, nargs='+', default=None,
        metavar='PROTEIN',
        help=f'Protein names to use (space-separated, case-insensitive). Default: {DEFAULT_PANEL}'
    )
    feat_grp.add_argument(
        '--protein-indices', type=int, nargs='+', default=None,
        metavar='IDX',
        help='0-based column indices (overrides name lookup). Legacy option.'
    )

    parser.add_argument('--n-neighbors', type=int,  default=5)
    parser.add_argument('--leaf-size',   type=int,  default=30)
    parser.add_argument('--algorithm',   type=str,  default='auto')
    parser.add_argument('--weights',     type=str,  default='distance')
    parser.add_argument('--p',           type=int,  default=2)

    parser.add_argument('--use-smote',    action='store_true', default=False)
    parser.add_argument('--n-iterations', type=int, default=100)

    parser.add_argument('--output-dir',  type=str, default='./results')
    parser.add_argument('--plot-curves', action='store_true', default=False)
    parser.add_argument('--verbose',     action='store_true', default=False)

    args = parser.parse_args()
    print_banner()
    os.makedirs(args.output_dir, exist_ok=True)
    start_time = time.time()

    if args.mode == 'internal':
        if not args.input:
            parser.error('--input is required for internal validation')
        internal_validation(args)
    elif args.mode == 'external':
        if not args.train_input or not args.test_input:
            parser.error('--train-input and --test-input are required')
        external_validation(args)

    print("\n" + "=" * 70)
    print("PreRisk-CoV2 completed successfully!")
    print(f"Total execution time : {time.time() - start_time:.2f} s")
    print("=" * 70 + "\n")


if __name__ == '__main__':
    main()
